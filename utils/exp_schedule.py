import requests
import os
import openpyxl
import json

dl_link = "https://cloud.hcu-hamburg.de/nextcloud/s/xkX3RHDqo3cEsti/download/deep_segmentation_experiments.xlsx"
dav_url = "https://cloud.hcu-hamburg.de/nextcloud/remote.php/dav/files/19A79C6B-3077-4EBF-A96C-3905CEFEFF0B/"

data_folder = "./data/"

def get_experiments_file(to_folder="./"):
    os.makedirs(to_folder, exist_ok=True)
    local_filename = to_folder + "/" + dl_link.split('/')[-1]
    # NOTE the stream=True parameter below
    with requests.get(dl_link, stream=True) as r:
        r.raise_for_status()
        with open(local_filename, 'wb') as f:
            for chunk in r.iter_content(chunk_size=8192): 
                # If you have chunk encoded response uncomment if
                # and set chunk_size parameter to None.
                #if chunk: 
                f.write(chunk)
    return local_filename

def get_creds(cred_path="credentials.json"):
    """ get user and pass from a json file"""
    with open(cred_path) as fr:
        creds = json.load(fr)
    return (creds["user"],creds["pass"])

def upload_experiments_file(filepath, remotepath="phd/testfile.xlsx"):
    """upload file at filepath to cloud with path+name given in remotepath via webdav PUT request"""
    print("uploading to cloud...")
    with open(filepath, 'rb') as f:
        r = requests.put(dav_url+"/"+remotepath, data=f, auth=get_creds())
        print(r)
    print("done uploading")

def write_experiment_to_excel(infile, exp):
    target_exp = exp["Exp #"]
    print("target",target_exp)
    my_wb = openpyxl.load_workbook(infile)
    ws = my_wb.active
    header = list(filter(lambda x: x,list(map(lambda x: x.value, next(ws.rows)))))
    for row in ws.iter_rows(min_row=2):
        if row[0].value == target_exp:
            print("updating experiments list...")
            for idx,cell in enumerate(zip(header,row)):
                col = header[idx]
                new_val = exp[col]
                print(cell[0],"old val",cell[1].value,"new val", new_val)
                cell[1].value = new_val
            my_wb.save(infile)
            print("done updating!")
            break

def get_next_experiment(experiments):
    for idx,e in enumerate(experiments[::-1]):
        if e["test dice"]:
            if idx > 0:
                return experiments[::-1][idx-1]
            else:
                return None

def get_all_experiments(infile):
    my_wb = openpyxl.load_workbook(infile)
    ws = my_wb.active
    header = list(map(lambda x: x.value, next(ws.rows)))
    e = []
    for row in ws.iter_rows(min_row=2):
        if not row[0].value:
            continue
        line = map(lambda x: x.value,row)
        e.append(dict(zip(header,line)))
    return e

if __name__ == "__main__":
    infile = get_experiments_file(data_folder)
    print(infile)
    exps = get_all_experiments(infile)
    print(exps)
    ne = get_next_experiment(exps)
    print(ne)
    ne["test dice"] = 1
    write_experiment_to_excel("data\deep_segmentation_experiments.xlsx",ne)
    upload_experiments_file(infile)